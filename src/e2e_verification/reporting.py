from __future__ import annotations

import html
import json
from pathlib import Path
from typing import Any
from urllib.parse import quote


def write_html_report(run_path: Path, output_path: Path | None = None) -> Path:
    state = json.loads(run_path.read_text(encoding="utf-8"))
    output_path = output_path or run_path.with_name("report.html")
    steps = state.get("steps", {})
    rows = "\n".join(_step_row(step_id, value) for step_id, value in steps.items())
    functional = _aggregate(steps, "functional_status", legacy=True)
    usability = _aggregate(steps, "usability_status", default="SKIP")
    case_counters = _case_counters(steps)
    gallery, warnings, filters = _gallery(steps, run_path.parent)
    findings = _findings(steps)
    warning_html = "" if not warnings else "<section class=warnings><h2>Artifact warnings</h2><ul>" + "".join(
        f"<li>{html.escape(item)}</li>" for item in warnings
    ) + "</ul></section>"
    document = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(str(state.get('workflow', 'Verification report')))}</title>
  <style>
    :root {{ color-scheme: light; font-family: Inter, ui-sans-serif, system-ui, sans-serif; color: #172033; background: #f5f7fb; }}
    * {{ box-sizing: border-box; }} body {{ max-width: 1200px; margin: 0 auto; padding: 2rem 1.25rem 4rem; line-height: 1.5; }}
    header, .verdicts, .filters {{ display: flex; flex-wrap: wrap; align-items: end; justify-content: space-between; gap: 1rem; }}
    .verdict {{ background: white; border: 1px solid #dbe1ea; border-radius: .75rem; padding: .8rem 1rem; min-width: 12rem; }}
    .counter {{ display: block; margin-top: .35rem; color: #536078; font-size: .85rem; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 2rem; background: white; }}
    th, td {{ padding: .75rem; border-bottom: 1px solid #dbe1ea; text-align: left; vertical-align: top; }}
    code {{ font-size: .9em; overflow-wrap: anywhere; }} .status {{ font-weight: 700; }}
    .gallery {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 1rem; margin-top: 1rem; }}
    figure {{ margin: 0; padding: .75rem; background: white; border: 1px solid #dbe1ea; border-radius: .75rem; }}
    figure img {{ display: block; width: 100%; height: 190px; object-fit: contain; background: #eef1f6; }}
    figcaption {{ padding-top: .6rem; overflow-wrap: anywhere; }} .meta {{ color: #536078; font-size: .85rem; }}
    .warnings {{ border: 1px solid #c88719; background: #fff8e8; padding: 0 1rem; margin-top: 1.5rem; }}
    select {{ min-height: 2.5rem; padding: .4rem; }}
  </style>
</head>
<body>
  <header><div><p>E2E verification</p><h1>{html.escape(str(state.get('workflow', 'Run')))}</h1></div><p class="status">{html.escape(str(state.get('status', 'UNKNOWN')))}</p></header>
  <div class="verdicts"><p class="verdict"><strong>Functional</strong><br>{html.escape(functional)}<span class="counter">{html.escape(_counter_text(case_counters, 'functional'))}</span></p><p class="verdict"><strong>Usability</strong><br>{html.escape(usability)}<span class="counter">{html.escape(_counter_text(case_counters, 'usability'))}</span></p></div>
  <p>Run <code>{html.escape(str(state.get('run_id', '')))}</code><br>Profile <code>{html.escape(str(state.get('profile', '')))}</code><br>Started <code>{html.escape(str(state.get('started_at', '')))}</code><br>Finished <code>{html.escape(str(state.get('finished_at', '')))}</code></p>
  <table><thead><tr><th>Step</th><th>Harness</th><th>Functional</th><th>Usability</th><th>Summary</th></tr></thead><tbody>{rows}</tbody></table>
  {warning_html}
  <section><h2>Findings</h2>{findings or '<p>No findings.</p>'}</section>
  <section><h2>Image evidence</h2>{filters}<div class="gallery">{gallery or '<p>No image artifacts.</p>'}</div></section>
  <script>for(const s of document.querySelectorAll('[data-filter]'))s.addEventListener('change',()=>{{for(const c of document.querySelectorAll('figure[data-shard]'))c.hidden=[...document.querySelectorAll('[data-filter]')].some(f=>f.value&&c.dataset[f.dataset.filter]!==f.value)}});</script>
</body>
</html>
"""
    output_path.write_text(document, encoding="utf-8")
    return output_path


def write_xlsx_report(run_path: Path, output_path: Path | None = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as error:
        raise RuntimeError("XLSX reports require: pip install 'e2e-verification[xlsx]'") from error
    state = json.loads(run_path.read_text(encoding="utf-8"))
    output_path = output_path or run_path.with_name("report.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Verification"
    for row in (["Workflow", state.get("workflow", "")], ["Run ID", state.get("run_id", "")],
                ["Profile", state.get("profile", "")], ["Status", state.get("status", "")],
                ["Started", state.get("started_at", "")], ["Finished", state.get("finished_at", "")], []):
        sheet.append(row)
    headers = ["Step", "Harness", "Risk", "Functional", "Usability", "Passed", "Failed", "Blocked", "Review"]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", bold=True); cell.fill = header_fill
    for step_id, value in state.get("steps", {}).items():
        summary = value.get("summary", {})
        sheet.append([step_id, value.get("harness", ""), value.get("risk", ""),
                      _functional(value), value.get("usability_status", "SKIP"),
                      summary.get("passed", 0), summary.get("failed", 0),
                      summary.get("blocked", 0), summary.get("review", 0)])
    sheet.freeze_panes = "A9"; sheet.auto_filter.ref = f"A8:I{sheet.max_row}"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
    workbook.save(output_path)
    return output_path


def _functional(value: dict[str, Any]) -> str:
    if value.get("functional_status"):
        return str(value["functional_status"])
    legacy = str(value.get("status", "BLOCKED"))
    return "PASS" if legacy == "REVIEW" else legacy


def _aggregate(steps: dict[str, dict[str, Any]], key: str, default: str = "BLOCKED", legacy: bool = False) -> str:
    values = [(_functional(item) if legacy else str(item.get(key, default))) for item in steps.values()]
    order = ["FAIL", "BLOCKED", "REVIEW", "PASS", "SKIP"]
    return next((status for status in order if status in values), default)


def _step_row(step_id: str, value: dict[str, Any]) -> str:
    summary = ", ".join(f"{key}: {item}" for key, item in value.get("summary", {}).items()) or "—"
    values = (step_id, value.get("harness", ""), _functional(value), value.get("usability_status", "SKIP"), summary)
    return "<tr>" + "".join(f"<td>{html.escape(str(item))}</td>" for item in values) + "</tr>"


def _case_counters(steps: dict[str, dict[str, Any]]) -> dict[str, dict[str, int] | int]:
    counters: dict[str, dict[str, int] | int] = {
        "total": 0,
        "functional": {status: 0 for status in ("PASS", "FAIL", "BLOCKED", "SKIP")},
        "usability": {status: 0 for status in ("PASS", "REVIEW", "BLOCKED", "SKIP")},
    }
    for step in steps.values():
        ui_audit = step.get("metadata", {}).get("ui_audit", {})
        cases = ui_audit.get("cases", []) if isinstance(ui_audit, dict) else []
        for case in cases if isinstance(cases, list) else []:
            if not isinstance(case, dict):
                continue
            counters["total"] = int(counters["total"]) + 1
            for axis in ("functional", "usability"):
                values = counters[axis]
                assert isinstance(values, dict)
                status = str(case.get(f"{axis}_status", "BLOCKED"))
                values[status if status in values else "BLOCKED"] += 1
    return counters


def _counter_text(counters: dict[str, dict[str, int] | int], axis: str) -> str:
    total = int(counters["total"])
    if not total:
        return "No UI cases"
    values = counters[axis]
    assert isinstance(values, dict)
    tail = ("FAIL", "BLOCKED", "SKIP") if axis == "functional" else ("REVIEW", "BLOCKED", "SKIP")
    return " · ".join([f"{values['PASS']}/{total} PASS", *(f"{status} {values[status]}" for status in tail)])


def _gallery(steps: dict[str, dict[str, Any]], root: Path) -> tuple[str, list[str], str]:
    cards: list[str] = []
    warnings: list[str] = []
    facets: dict[str, set[str]] = {"role": set(), "state": set(), "viewport": set(), "page": set(), "shard": set()}
    resolved_root = root.resolve()
    for step_id, step in steps.items():
        for artifact in step.get("artifacts", []):
            raw = str(artifact.get("path", ""))
            kind = str(artifact.get("kind", ""))
            if kind not in {"screenshot", "image"} and Path(raw).suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp"}:
                continue
            path = Path(raw)
            if not raw or path.is_absolute():
                warnings.append(f"{step_id}: unsafe image artifact path: {raw or '(empty)'}")
                continue
            candidate = (resolved_root / path).resolve()
            try:
                candidate.relative_to(resolved_root)
            except ValueError:
                warnings.append(f"{step_id}: image artifact escapes run directory: {raw}")
                continue
            if not candidate.is_file():
                warnings.append(f"{step_id}: image artifact is missing: {raw}")
                continue
            href = quote(path.as_posix(), safe="/-._~")
            label = artifact.get("description") or artifact.get("case_id") or path.name
            meta = " · ".join(str(artifact.get(key, "")) for key in ("page", "role", "state", "variant") if artifact.get(key))
            viewport = artifact.get("viewport", {})
            viewport_name = str(viewport.get("name", "")) if isinstance(viewport, dict) else str(viewport)
            values = {
                "role": str(artifact.get("role", "")),
                "state": str(artifact.get("state", "")),
                "viewport": viewport_name,
                "page": str(artifact.get("page", "")),
                "shard": str(artifact.get("shard", "") or step_id),
            }
            for key, value in values.items():
                if value: facets[key].add(value)
            cards.append(
                f'<figure data-role="{html.escape(values["role"], quote=True)}" data-state="{html.escape(values["state"], quote=True)}" data-viewport="{html.escape(values["viewport"], quote=True)}" data-page="{html.escape(values["page"], quote=True)}" data-shard="{html.escape(values["shard"], quote=True)}"><a href="{html.escape(href, quote=True)}"><img loading="lazy" src="{html.escape(href, quote=True)}" alt="{html.escape(str(label), quote=True)}"></a>'
                f'<figcaption><strong>{html.escape(str(label))}</strong><div class="meta">{html.escape(meta)}</div></figcaption></figure>'
            )
    selectors = []
    for key, values in facets.items():
        options = '<option value="">All</option>' + "".join(
            f'<option value="{html.escape(value, quote=True)}">{html.escape(value)}</option>' for value in sorted(values)
        )
        selectors.append(f'<label>{key.title()} <select data-filter="{key}">{options}</select></label>')
    return "".join(cards), warnings, '<div class="filters">' + "".join(selectors) + "</div>"


def _findings(steps: dict[str, dict[str, Any]]) -> str:
    rows: list[str] = []
    for step_id, step in steps.items():
        for finding in step.get("findings", []):
            evidence = ", ".join(str(item) for item in finding.get("evidence", []))
            rows.append("<tr>" + "".join(f"<td>{html.escape(str(value))}</td>" for value in (
                step_id, finding.get("severity", ""), finding.get("category", ""),
                finding.get("status", ""), finding.get("title", ""), evidence,
            )) + "</tr>")
    if not rows:
        return ""
    return '<table><thead><tr><th>Step</th><th>Severity</th><th>Category</th><th>Status</th><th>Finding</th><th>Evidence</th></tr></thead><tbody>' + "".join(rows) + "</tbody></table>"
